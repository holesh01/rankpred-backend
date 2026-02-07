from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import shutil
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import requests
import warnings
warnings.filterwarnings("ignore", category=Warning)
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from flask import send_file
import io

app = Flask(__name__)
CORS(app)

BASE_DIR = "data"
UPLOAD_DIR = "uploads"

os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ---------------- BASIC ----------------
@app.route("/")
def home():
    return "Backend running"


def safe_name(name):
    return name.replace(" ", "_")

def make_shift_id(exam_date, exam_time):
    """
    exam_date: 13/10/2025
    exam_time: 9:30 AM - 11:00 AM
    """
    date_part = exam_date.replace("/", "-")          # 13-10-2025
    time_part = exam_time.replace(":", "-")           # 9-30 AM - 11-00 AM
    time_part = re.sub(r"\s+", "", time_part)         # remove spaces

    return f"{date_part}_{time_part}"


# ---------------- ADMIN: LIST EXAMS ----------------
@app.route("/admin/exams", methods=["GET"])
def admin_list_exams():
    exams = []

    for folder in os.listdir(BASE_DIR):
        exam_dir = os.path.join(BASE_DIR, folder)
        scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")

        if not os.path.isdir(exam_dir) or not os.path.exists(scheme_path):
            continue

        wb = load_workbook(scheme_path)
        ws = wb.active

        scheme = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            scheme[row[0]] = row[1]

        exams.append({
            "exam_name": scheme.get("Exam Name"),
            "correct": scheme.get("Correct"),
            "wrong": scheme.get("Wrong"),
            "na": scheme.get("NA")
        })

    return jsonify(exams)


# ---------------- ADMIN: CREATE EXAM ----------------
@app.route("/admin/create-exam", methods=["POST"])
def create_exam():
    data = request.json

    exam_name = data.get("exam_name")
    correct = data.get("correct")
    wrong = data.get("wrong")
    na = data.get("na")
    subjects = data.get("subjects", [])

    if not exam_name:
        return jsonify({"error": "exam_name required"}), 400

    if not subjects:
        return jsonify({"error": "subjects required"}), 400

    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    os.makedirs(exam_dir, exist_ok=True)

    # ---- marking_scheme.xlsx ----
    scheme_path = os.path.join(exam_dir, "marking_scheme.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "scheme"
    ws.append(["Key", "Value"])
    ws.append(["Exam Name", exam_name])
    ws.append(["Correct", correct])
    ws.append(["Wrong", wrong])
    ws.append(["NA", na])
    wb.save(scheme_path)

    # ---- subjects.xlsx ----
    subjects_path = os.path.join(exam_dir, "subjects.xlsx")
    wb_sub = Workbook()
    ws_sub = wb_sub.active
    ws_sub.title = "subjects"
    ws_sub.append(["Subject Name", "Max Marks", "Count In Total"])

    for sub in subjects:
        ws_sub.append([
            sub["name"],
            sub["max_marks"],
            "YES" if sub["count_in_total"] else "NO"
        ])

    wb_sub.save(subjects_path)

    # ---- responses.xlsx (TEMPLATE) ----
    response_path = os.path.join(exam_dir, "responses.xlsx")

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "responses"

    header = ["Name", "Roll", "Category", "Gender", "State", "Final Marks"]

    for sub in subjects:
        base = sub["name"]
        header.extend([
            f"{base}_Attempt",
            f"{base}_R",
            f"{base}_W",
            f"{base}_NA",
            f"{base}_Marks"
        ])

    ws2.append(header)
    wb2.save(response_path)

    return jsonify({"status": "success", "exam": exam_name})


# ---------------- ADMIN: DELETE EXAM ----------------
@app.route("/admin/delete-exam", methods=["POST"])
def delete_exam():
    data = request.json
    exam_name = data.get("exam_name")

    if not exam_name:
        return jsonify({"error": "exam_name required"}), 400

    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    if not os.path.exists(exam_dir):
        return jsonify({"error": "exam not found"}), 404

    shutil.rmtree(exam_dir)
    return jsonify({"status": "deleted"})


# ---------------- READ MARKING SCHEME ----------------
def read_marking_scheme(exam_name):
    path = os.path.join(BASE_DIR, safe_name(exam_name), "marking_scheme.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    scheme = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        scheme[row[0]] = row[1]
    return scheme


# ---------------- READ SUBJECT CONFIG ----------------
def read_subjects(exam_name):
    path = os.path.join(BASE_DIR, safe_name(exam_name), "subjects.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    subjects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        subjects.append({
            "name": row[0],
            "max_marks": row[1],
            "count_in_total": True if row[2] == "YES" else False
        })
    return subjects


def parse_response_sectionwise_from_html(html_content, scheme, subjects):
    section_map = {}
    section_order = []
    current_section = None

    soup = BeautifulSoup(html_content, "lxml")

    for el in soup.find_all("div"):
        if el.get("class") == ["section-lbl"]:
            current_section = el.get_text(strip=True)
            if current_section not in section_map:
                section_map[current_section] = {"c": 0, "w": 0, "n": 0}
                section_order.append(current_section)

        if el.get("class") == ["question-pnl"] and current_section:
            chosen = None
            correct_option = None

            chosen_row = el.find("td", string=re.compile("Chosen Option"))
            if chosen_row:
                v = chosen_row.find_next_sibling("td").get_text(strip=True)
                if v != "--":
                    chosen = int(v)

            right_td = el.find("td", class_="rightAns")
            if right_td:
                m = re.search(r"(\d)\.", right_td.get_text())
                if m:
                    correct_option = int(m.group(1))

            if chosen is None:
                section_map[current_section]["n"] += 1
            elif chosen == correct_option:
                section_map[current_section]["c"] += 1
            else:
                section_map[current_section]["w"] += 1

    subject_results = []
    final_marks = 0

    for idx, sec in enumerate(section_order):
        c = section_map[sec]["c"]
        w = section_map[sec]["w"]
        n = section_map[sec]["n"]

        marks = (
            c * scheme["Correct"]
            + w * scheme["Wrong"]
            + n * scheme["NA"]
        )

        sub = subjects[idx]

        subject_results.append({
            "name": sub["name"],
            "attempt": c + w,
            "r": c,
            "w": w,
            "na": n,
            "marks": marks,
            "count_in_total": sub["count_in_total"]
        })


        if sub["count_in_total"]:
            final_marks += marks

    return final_marks, subject_results

# ---------------- SAVE RESULT ----------------
def save_user_result(exam_name, shift_id, base_data, subject_results):
    exam_dir = os.path.join(BASE_DIR, safe_name(exam_name))
    path = os.path.join(exam_dir, f"responses_{shift_id}.xlsx")

    # create shift file from template if not exists
    if not os.path.exists(path):
        template = os.path.join(exam_dir, "responses.xlsx")
        shutil.copy(template, path)

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {h: i for i, h in enumerate(headers)}

    roll_value = base_data[col_index["Roll"]]
    updated = False

    for row_idx in range(2, ws.max_row + 1):
        existing_roll = ws.cell(row=row_idx, column=col_index["Roll"] + 1).value
        if str(existing_roll) == str(roll_value):
            for key in ["Name", "Category", "Gender", "State", "Final Marks"]:
                ws.cell(row=row_idx, column=col_index[key] + 1,
                        value=base_data[col_index[key]])

            for sub in subject_results:
                base = sub["name"]
                ws.cell(row=row_idx, column=col_index[f"{base}_Attempt"] + 1, value=sub["attempt"])
                ws.cell(row=row_idx, column=col_index[f"{base}_R"] + 1, value=sub["r"])
                ws.cell(row=row_idx, column=col_index[f"{base}_W"] + 1, value=sub["w"])
                ws.cell(row=row_idx, column=col_index[f"{base}_NA"] + 1, value=sub["na"])
                ws.cell(row=row_idx, column=col_index[f"{base}_Marks"] + 1, value=sub["marks"])

            updated = True
            break

    if not updated:
        row = base_data[:]
        for sub in subject_results:
            row.extend([sub["attempt"], sub["r"], sub["w"], sub["na"], sub["marks"]])
        ws.append(row)

    wb.save(path)
    calculate_shift_ranks(path)



#  ---------------- EVALUATE ----------------
@app.route("/evaluate", methods=["POST"])
def evaluate_exam():
    return jsonify({
        "error": "File upload not supported. Please use DigiALM link."
    }), 400

# @app.route("/evaluate", methods=["POST"])
# def evaluate_exam():
#     exam_name = request.form.get("exam_name")
#     name = request.form.get("name")
#     roll = request.form.get("roll")
#     category = request.form.get("category")
#     gender = request.form.get("gender")
#     state = request.form.get("state")
#     file = request.files.get("file")

#     if not all([exam_name, name, roll, category, gender, state, file]):
#         return jsonify({"error": "Missing fields"}), 400

#     upload_path = os.path.join(UPLOAD_DIR, file.filename)
#     file.save(upload_path)

#     scheme = read_marking_scheme(exam_name)
#     subjects = read_subjects(exam_name)

#     with open(upload_path, "r", encoding="utf-8", errors="ignore") as f:html_content = f.read()

#     final_marks, subject_results = parse_response_sectionwise_from_html(html_content, scheme, subjects
# )


#     save_user_result(
#         exam_name,
#         [name, roll, category, gender, state, final_marks],
#         subject_results
#     )

#     return jsonify({"status": "saved", "final_marks": final_marks})

def extract_candidate_details(html_content):
    soup = BeautifulSoup(html_content, "lxml")
    text = soup.get_text(separator=" ")

    details = {
        "roll": None,
        "name": None,
        "venue": None,
        "exam_date": None,
        "exam_time": None,
        "subject": None
    }

    # -------- TABLE-BASED DETAILS (STRICT) --------
    for row in soup.select("table tr"):
        cells = [c.get_text(strip=True) for c in row.find_all("td")]
        if len(cells) != 2:
            continue

        key, value = cells
        key_lower = key.lower()

        if "roll" in key_lower:
            details["roll"] = value

        elif "candidate name" in key_lower:
            details["name"] = value   # ✅ ONLY candidate name

        elif "venue" in key_lower or "test center" in key_lower or "centre" in key_lower:
            details["venue"] = value

        elif "subject" in key_lower:
            details["subject"] = value

    # -------- DATE EXTRACTION (REGEX) --------
    date_match = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", text)
    if date_match:
        details["exam_date"] = date_match.group(1)

    # -------- TIME EXTRACTION (REGEX) --------
    time_match = re.search(
        r"(\d{1,2}:\d{2}\s*(?:AM|PM)\s*-\s*\d{1,2}:\d{2}\s*(?:AM|PM))",
        text,
        re.IGNORECASE
    )
    if time_match:
        details["exam_time"] = time_match.group(1).upper()

    return details


@app.route("/evaluate-from-url", methods=["POST"])
def evaluate_exam_from_url():
    data = request.json

    exam_name = data.get("exam_name")
    category = data.get("category")
    gender = data.get("gender")
    state = data.get("state")
    url = data.get("url")

    if not all([exam_name, category, gender, state, url]):
        return jsonify({"error": "Missing fields"}), 400

    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, timeout=20)

        if response.status_code != 200:
            return jsonify({"error": "Unable to fetch DigiALM page"}), 400

        html_content = response.text

        # 🔥 AUTO-EXTRACT DETAILS
        details = extract_candidate_details(html_content)
        

        if not details["exam_date"] or not details["exam_time"]:
            return jsonify({
                "error": "Exam date/time not found in response sheet. Please use URL method."
            }), 400

        shift_id = make_shift_id(details["exam_date"], details["exam_time"])


        if not details["roll"] or not details["name"]:
            return jsonify({"error": "Unable to extract candidate details"}), 400

        scheme = read_marking_scheme(exam_name)
        subjects = read_subjects(exam_name)

        final_marks, subject_results = parse_response_sectionwise_from_html(
            html_content, scheme, subjects
        )

        save_user_result(
            exam_name,
            shift_id,
            [
                details["name"],
                details["roll"],
                category,
                gender,
                state,
                final_marks
            ],
            subject_results
        )


        return jsonify({
            "status": "saved",
            "final_marks": final_marks,
            "candidate": details
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------- RESULT API ----------------
@app.route("/result")
def get_result():
    exam = request.args.get("exam")
    roll = request.args.get("roll")

    if not exam or not roll:
        return jsonify({"error": "Missing exam or roll"}), 400

    exam_dir = os.path.join(BASE_DIR, safe_name(exam))
    if not os.path.exists(exam_dir):
        return jsonify({"error": "Exam not found"}), 404

    subjects_cfg = read_subjects(exam)

    # 🔍 search candidate in all shift files
    for file in os.listdir(exam_dir):
        if not file.startswith("responses_") or not file.endswith(".xlsx"):
            continue

        path = os.path.join(exam_dir, file)
        wb = load_workbook(path)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        # total candidates in this shift
        total_candidates = ws.max_row - 1  # excluding header

        for r in ws.iter_rows(min_row=2, values_only=True):
            if str(r[col["Roll"]]) == str(roll):

                shift_id = file.replace("responses_", "").replace(".xlsx", "")

                counted = []
                qualifying = []
                final_marks = 0

                # ---------- SUBJECT-WISE PERFORMANCE ----------
                for sub in subjects_cfg:
                    base = sub["name"]
                    item = {
                        "name": base,
                        "attempt": r[col[f"{base}_Attempt"]],
                        "na": r[col[f"{base}_NA"]],
                        "r": r[col[f"{base}_R"]],
                        "w": r[col[f"{base}_W"]],
                        "marks": r[col[f"{base}_Marks"]]
                    }

                    if sub["count_in_total"]:
                        counted.append(item)
                        final_marks += item["marks"]
                    else:
                        qualifying.append(item)

                # ---------- FINAL RESPONSE ----------
                return jsonify({
                    "exam": exam,
                    "shift_id": shift_id,
                    "candidate": {
                        "name": r[col["Name"]],
                        "roll": r[col["Roll"]],
                        "category": r[col["Category"]],
                        "gender": r[col["Gender"]],
                        "state": r[col["State"]],
                        "rank": r[col.get("Rank")]
                    },
                    "total_candidates": total_candidates,
                    "counted_subjects": counted,
                    "qualifying_subjects": qualifying,
                    "final_marks": final_marks
                })

    return jsonify({"error": "Candidate not found"}), 404


@app.route("/result-pdf")
def result_pdf():
    exam = request.args.get("exam")
    roll = request.args.get("roll")

    if not exam or not roll:
        return jsonify({"error": "Missing exam or roll"}), 400

    # reuse existing result logic
    result = get_result().get_json()
    if "error" in result:
        return jsonify(result), 400

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ---------- WATERMARK ----------
    c.saveState()
    c.setFont("Helvetica-Bold", 60)
    c.setFillGray(0.9)
    c.translate(width / 2, height / 2)
    c.rotate(45)
    c.drawCentredString(0, 0, "RankPred")
    c.restoreState()

    y = height - 50

    # ---------- HEADER ----------
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y, result["exam"])
    y -= 30

    # ---------- CANDIDATE DETAILS ----------
    c.setFont("Helvetica", 11)
    cand = result["candidate"]

    details = [
        f"Name: {cand['name']}",
        f"Roll No: {cand['roll']}",
        f"Category: {cand['category']}",
        f"State: {cand['state']}"
    ]

    for d in details:
        c.drawString(50, y, d)
        y -= 15

    y -= 15

    # ---------- TABLE CONFIG ----------
    table_x = 40
    col_widths = [110, 55, 45, 40, 40, 55]
    headers = ["Subject", "Attempt", "NA", "R", "W", "Marks"]
    row_height = 18

    def draw_row(values, y, bold=False, bg=False):
        x = table_x
        if bg:
            c.setFillGray(0.9)
            c.rect(x, y - row_height + 3, sum(col_widths), row_height, fill=1, stroke=0)
            c.setFillGray(0)

        c.setFont("Helvetica-Bold" if bold else "Helvetica", 10)

        for i, v in enumerate(values):
            c.rect(x, y - row_height + 3, col_widths[i], row_height, stroke=1, fill=0)
            c.drawCentredString(x + col_widths[i] / 2, y - 12, str(v))
            x += col_widths[i]

    # ---------- TABLE HEADER ----------
    draw_row(headers, y, bold=True, bg=True)
    y -= row_height

    # ---------- TABLE ROWS ----------
    for sub in result["counted_subjects"] + result["qualifying_subjects"]:
        draw_row(
            [
                sub["name"],
                sub["attempt"],
                sub["na"],
                sub["r"],
                sub["w"],
                sub["marks"]
            ],
            y
        )
        y -= row_height

        if y < 80:
            c.showPage()
            y = height - 50

    y -= 20

    # ---------- FINAL MARKS ----------
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, f"Final Marks: {result['final_marks']}")
    y -= 20

    # ---------- SHIFT RANK ----------
    rank = cand.get("rank", "-")
    total = result.get("total_candidates", "-")

    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, f"Shift Rank: {rank} / {total}")

    c.showPage()
    c.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{cand['roll']}_result.pdf",
        mimetype="application/pdf"
    )

def calculate_shift_ranks(path):
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    # add Rank column if not present
    if "Rank" not in col:
        ws.cell(row=1, column=len(headers) + 1, value="Rank")
        col["Rank"] = len(headers)

    # collect (row_index, marks)
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        marks = ws.cell(row=row_idx, column=col["Final Marks"] + 1).value
        if marks is not None:
            rows.append((row_idx, marks))

    # sort by marks descending
    rows.sort(key=lambda x: x[1], reverse=True)

    current_rank = 1
    prev_marks = None
    processed = 0

    for i, (row_idx, marks) in enumerate(rows):
        if marks != prev_marks:
            current_rank = processed + 1

        ws.cell(row=row_idx, column=col["Rank"] + 1, value=current_rank)

        prev_marks = marks
        processed += 1

    wb.save(path)
@app.route("/evaluate-html", methods=["POST"])
def evaluate_exam_from_html():
    data = request.json

    exam_name = data.get("exam_name")
    category = data.get("category")
    gender = data.get("gender")
    state = data.get("state")
    html_content = data.get("html")

    if not all([exam_name, category, gender, state, html_content]):
        return jsonify({"error": "Missing fields"}), 400

    # extract candidate details
    details = extract_candidate_details(html_content)

    if not details["roll"] or not details["name"]:
        return jsonify({"error": "Unable to extract candidate details"}), 400

    if not details["exam_date"] or not details["exam_time"]:
        return jsonify({"error": "Exam date/time not found"}), 400

    shift_id = make_shift_id(details["exam_date"], details["exam_time"])

    scheme = read_marking_scheme(exam_name)
    subjects = read_subjects(exam_name)

    final_marks, subject_results = parse_response_sectionwise_from_html(
        html_content, scheme, subjects
    )

    save_user_result(
        exam_name,
        shift_id,
        [
            details["name"],
            details["roll"],
            category,
            gender,
            state,
            final_marks
        ],
        subject_results
    )

    return jsonify({
        "status": "saved",
        "final_marks": final_marks,
        "roll": details["roll"]
    })


# ---------------- START ----------------
# if __name__ == "__main__":
#     app.run(debug=True)
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
